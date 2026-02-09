from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def write_to_excel(results, output_path):
    wb = Workbook()
    
    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF")
    # हेडर के लिए गहरा नीला रंग
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    # अल्टरनेटिंग कलर्स (एक सफेद, एक हल्का नीला/ग्रे)
    color_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    color_even = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color="B2B2B2"),
        right=Side(style='thin', color="B2B2B2"),
        top=Side(style='thin', color="B2B2B2"),
        bottom=Side(style='thin', color="B2B2B2")
    )
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Invoice Summary"
    headers1 = ["File Name", "Invoice No", "Date", "Supplier Name", "Supplier GSTIN", "Buyer Name", "Buyer GSTIN", "Total Amount"]
    ws1.append(headers1)
    
    # Header Styling
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data Rows with Alternating Colors
    for i, res in enumerate(results, start=2):
        row_data = [
            res.get('file'), res.get('invoice_no'), res.get('invoice_date'),
            res.get('supplier_name'), res.get('supplier_gstin'),
            res.get('buyer_name'), res.get('buyer_gstin'),
            res.get('total_amount')
        ]
        ws1.append(row_data)
        
        current_fill = color_even if i % 2 == 0 else color_odd
        for cell in ws1[i]:
            cell.fill = current_fill
            cell.border = thin_border
            cell.alignment = left_align

    # --- Sheet 2: Item Details ---
    ws2 = wb.create_sheet("Item Details")
    headers2 = ["File Name", "Item Description", "Qty", "Unit Rate", "GST %", "Amount"]
    ws2.append(headers2)
    
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid") # गहरा हरा
        cell.alignment = center_align

    row_idx = 2
    # इन्वेंट्री में हम फाइल के हिसाब से कलर बदलेंगे
    for file_idx, res in enumerate(results):
        current_fill = color_even if file_idx % 2 == 0 else color_odd
        
        for item in res.get('inventories', []):
            ws2.append([
                res.get('file'), item.get('item'), item.get('qty'),
                item.get('rate'), item.get('tax_rate'), item.get('amount')
            ])
            
            for cell in ws2[row_idx]:
                cell.fill = current_fill
                cell.border = thin_border
                cell.alignment = left_align
            row_idx += 1

    # --- Column Width Adjustments ---
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(output_path)
