import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def write_excel(results, all_debug, MODE, CFG):
    output_dir = "output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # ---------- Styles ----------
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    even_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ---------- 1. INVOICE SUMMARY ----------
    if CFG.get("invoice_summary"):
        ws_sum = wb.create_sheet("Invoice Summary")
        headers = [
            "File Name", "Invoice No", "Invoice Date",
            "Supplier Name", "Buyer Name", "Buyer GSTIN", "Total Amount"
        ]
        ws_sum.append(headers)

        for cell in ws_sum[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, r in enumerate(results, start=2):
            ws_sum.append([
                r["file"],
                r["invoice_no"],
                r["invoice_date"],
                r["supplier_name"],
                r["buyer_name"],
                r["buyer_gstin"],
                r["total_amount"]
            ])
            fill = even_fill if i % 2 == 0 else PatternFill(fill_type=None)
            for cell in ws_sum[i]:
                cell.fill = fill
                cell.border = border

    # ---------- 2. ITEM DETAILS ----------
    if CFG.get("item_details"):
        ws_item = wb.create_sheet("Item Details")
        ws_item.append(["File Ref", "Item Description", "Qty", "Rate", "Amount"])

        for cell in ws_item[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row_idx = 2
        for f_idx, r in enumerate(results):
            group_fill = even_fill if f_idx % 2 == 0 else PatternFill(fill_type=None)
            for item in r.get("inventories", []):
                ws_item.append([
                    r["file"],
                    item.get("item", ""),
                    item.get("qty", ""),
                    item.get("rate", ""),
                    item.get("amount", "")
                ])
                for cell in ws_item[row_idx]:
                    cell.fill = group_fill
                    cell.border = border
                row_idx += 1

    # ---------- 3. TALLY SALES ----------
    if CFG.get("tally_sales"):
        ws_tally = wb.create_sheet("Tally Sales")
        ws_tally.append(["Date", "Particulars", "Vch Type", "Vch No", "Debit", "Credit"])

        for cell in ws_tally[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in results:
            amount = r.get("total_amount", 0)

            ws_tally.append([
               r.get("invoice_date", ""),
               r.get("buyer_name", ""),
               "Sales",
               r.get("invoice_no", ""),
               None,        # Debit ALWAYS blank
               amount       # Credit ONLY
               ])

    # ---------- 4. MISSING CORRECTION ----------
    if CFG.get("missing_correction"):
        ws_miss = wb.create_sheet("Missing Correction")
        ws_miss.append(["File", "Field", "Observation"])

        for cell in ws_miss[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        if all_debug:
            for d in all_debug:
                ws_miss.append([d["file"], d["field"], d["value"]])
        else:
            ws_miss.append(["All Files", "Global", "No Critical Errors Found"])

    # ---------- 5. DASHBOARD ----------
    if CFG.get("dashboard"):
        ws_dash = wb.create_sheet("Dashboard")
        ws_dash.append(["REPORT STATISTICS", "VALUE"])

        for cell in ws_dash[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        total = len(results)
        auto = len([
            r for r in results
            if r["invoice_no"] != "N/A"
            and float(str(r["total_amount"]).replace(",", "")) > 0
        ])

        ws_dash.append(["Total Invoices Processed", total])
        ws_dash.append(["Status: AUTO (Success)", auto])
        ws_dash.append(["Status: REVIEW (Manual)", total - auto])
        ws_dash.append(["MODE", MODE])

        # 🔐 VERY IMPORTANT FIX (NO DICT TO EXCEL)
        for k, v in CFG.items():
            ws_dash.append([k, str(v)])

    # ---------- Final Formatting ----------
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    file_path = os.path.join(output_dir, f"GST_Final_Report_{MODE}.xlsx")
    wb.save(file_path)
    return file_path
