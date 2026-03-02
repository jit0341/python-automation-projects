#!/usr/bin/env python3
"""
🔥 SECL MASTER BILL - ULTRA CLEAN VERSION 🔥
Termux 100% Compatible | No Syntax Errors
"""

import json
import os
import re
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from num2words import num2words
    print("✅ All libraries imported")
except ImportError as e:
    print(f"❌ Missing library: {e}")
    print("📦 Install: pip install openpyxl num2words")
    exit(1)

# =====================================================
# CONFIG
# =====================================================
JSON_FILE = "march_2019.json"
TEMPLATE_FILE = "SECL_COMPLETE_TRUE_COPY_A4.xlsx"
OUTPUT_FILE = "output_filled.xlsx"

print("🚀 Starting...")

# File checks
if not os.path.exists(JSON_FILE):
    print(f"❌ {JSON_FILE} missing!")
    exit(1)
if not os.path.exists(TEMPLATE_FILE):
    print(f"❌ {TEMPLATE_FILE} missing!")
    exit(1)

print("✅ Files OK")

# =====================================================
# HELPERS
# =====================================================
def clean_number(value):
    if not value: return 0.0
    value = str(value).replace("Te", "").replace(",", "").strip()
    match = re.search(r"[-+]?d*.?d+", value)
    return float(match.group()) if match else 0.0

def normalize(text):
    return str(text).lower().replace("
", " ").strip()

def safe_write(sheet, cell_ref, value):
    cell = sheet[cell_ref]
    for merged in sheet.merged_cells.ranges:
        if cell.coordinate in merged:
            cell = sheet[merged.coord.split(":")[0]]
            break
    cell.value = value

def extract_table(response, table_idx):
    blocks = response["Blocks"]
    block_map = {b["Id"]: b for b in blocks}
    tables = [b for b in blocks if b["BlockType"] == "TABLE"]
    table = tables[table_idx]
    
    data = {}
    for rel in table.get("Relationships", []):
        if rel["Type"] == "CHILD":
            for cid in rel["Ids"]:
                cell = block_map[cid]
                if cell["BlockType"] == "CELL":
                    r = cell["RowIndex"]
                    c = cell["ColumnIndex"]
                    text = ""
                    for crel in cell.get("Relationships", []):
                        if crel["Type"] == "CHILD":
                            for wid in crel["Ids"]:
                                word = block_map[wid]
                                if word["BlockType"] == "WORD":
                                    text += word["Text"] + " "
                    if r not in data: data[r] = {}
                    data[r][c] = text.strip()
    return data

# =====================================================
# LOAD DATA
# =====================================================
print("📄 Loading JSON...")
with open(JSON_FILE, "r", encoding="utf-8") as f:
    response = json.load(f)

print("📊 Loading Excel...")
wb = load_workbook(TEMPLATE_FILE)
sheet = wb.active
tables = [b for b in response["Blocks"] if b["BlockType"] == "TABLE"]

# =====================================================
# FILL MAIN TABLE
# =====================================================
print("🔍 Filling data...")
for idx in range(len(tables)):
    table_data = extract_table(response, idx)
    
    for row_idx in table_data:
        row = table_data[row_idx]
        desc = normalize(row.get(2, ""))
        if "description" in desc: continue
        
        # Row mapping
        row_map = {
            "pay loader": 16,
            "western patch": 17,
            "central patch": 18,
            "eastern patch": 19
        }
        diesel_row = 21 if "transport" in desc else 22 if "loading" in desc else 0
        if "diesel escalation" in desc and diesel_row:
            excel_row = diesel_row
        else:
            excel_row = next((v for k, v in row_map.items() if k in desc), None)
            if not excel_row: continue
        
        # Fill columns C-I
        for col_idx, col_letter in zip([3,4,5,6,7,8,9], ["C","D","E","F","G","H","I"]):
            safe_write(sheet, f"{col_letter}{excel_row}", clean_number(row.get(col_idx)))

print("✅ Main table done")

# =====================================================
# SUBTOTAL
# =====================================================
sub_last = sum(float(sheet[f"G{r}"].value or 0) for r in [16,17,18,19])
sub_since = sum(float(sheet[f"H{r}"].value or 0) for r in [16,17,18,19])
sub_total = sum(float(sheet[f"I{r}"].value or 0) for r in [16,17,18,19])

safe_write(sheet, "G23", sub_last)
safe_write(sheet, "H23", sub_since)
safe_write(sheet, "I23", sub_total)

# =====================================================
# GST, DEDUCTIONS, SUMMARY (Combined)
# =====================================================
print("📊 Processing GST/Deductions/Summary...")
for idx in range(len(tables)):
    table_data = extract_table(response, idx)
    
    for row in table_data.values():
        row_text = normalize(" ".join(row.values()))
        
        # GST
        if "cgst" in row_text:
            cgst1, cgst2, cgst3 = [clean_number(row.get(c)) for c in [7,8,9]]
            safe_write(sheet, "F24", "CGST (9%)")
            safe_write(sheet, "G24", cgst1); safe_write(sheet, "H24", cgst2); safe_write(sheet, "I24", cgst3)
            
            # SGST
            for r2 in table_data.values():
                if "sgst" in normalize(" ".join(r2.values())):
                    safe_write(sheet, "F25", "SGST (9%)")
                    safe_write(sheet, "G25", clean_number(r2.get(7)))
                    safe_write(sheet, "H25", clean_number(r2.get(8)))
                    safe_write(sheet, "I25", clean_number(r2.get(9)))
                    break
        
        # Deductions
        if "i tax" in row_text: safe_write(sheet, "B28", clean_number(row.get(2)))
        elif "s.d" in row_text: safe_write(sheet, "B29", clean_number(row.get(2)))
        elif "electricity" in row_text: safe_write(sheet, "B30", clean_number(row.get(2)))
        elif "gst on tds" in row_text:
            target = "B31" if not sheet["B31"].value else "B32"
            safe_write(sheet, target, clean_number(row.get(2)))
        elif row_text.startswith("total"): safe_write(sheet, "B33", clean_number(row.get(2)))
        
        # Summary
        if "total upto" in row_text: safe_write(sheet, "I27", clean_number(row.get(2)))
        elif "paid" in row_text: safe_write(sheet, "I28", clean_number(row.get(2)))
        elif "since bill" in row_text: safe_write(sheet, "I29", clean_number(row.get(2)))
        elif "r/off" in row_text: safe_write(sheet, "I30", clean_number(row.get(2)))
        elif "deduction" in row_text: safe_write(sheet, "I31", clean_number(row.get(2)))

# =====================================================
# TOTAL WORDS
# =====================================================
total_val = sheet["I32"].value
if total_val:
    words = num2words(int(round(total_val))).replace(",", "").title()
    safe_write(sheet, "F35", f"Rupees {words} Only")
    sheet["F35"].alignment = Alignment(wrap_text=True)

# =====================================================
# LAYOUT (SIMPLE VERSION)
# =====================================================
print("🎨 Applying layout...")

# Header
sheet.merge_cells("A1:I1")
sheet["A1"].value = "South Eastern Coalfields Limited Bhatgaon Area"
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
sheet["A1"].font = Font(size=14, bold=True)
sheet.row_dimensions[1].height = 28

sheet.merge_cells("B3:I3")
sheet["B3"].alignment = Alignment(vertical="center")

sheet["B4"].value = "Name of Work:"
sheet["B4"].font = Font(bold=True)
sheet.merge_cells("B5:I6")
sheet["B5"].alignment = Alignment(wrap_text=True, vertical="top")

# Row heights
for r in range(2,12): sheet.row_dimensions[r].height = 20

# Columns
widths = {"A":6,"B":42,"C":12,"D":12,"E":12,"F":12,"G":14,"H":14,"I":16}
for col, w in widths.items():
    sheet.column_dimensions[col].width = w

print("✅ Layout done")

# =====================================================
# SAVE
# =====================================================
wb.save(OUTPUT_FILE)
print(f"
🎉 SUCCESS! Saved: {OUTPUT_FILE}")
print("📱 View: libreoffice output_filled.xlsx")
print("📱 Copy: cp output_filled.xlsx /sdcard/")
